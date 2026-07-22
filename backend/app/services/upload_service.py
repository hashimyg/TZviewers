import io
import re
import csv
import uuid
import logging
from typing import List, Dict, Any
from fastapi import UploadFile, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from openpyxl import load_workbook

from app.database.models.upload import Upload
from app.database.models.contact import Contact
from app.exceptions.base import ContactProcessingError

logger = logging.getLogger("app.upload_service")

MAX_UPLOAD_SIZE = 25 * 1024 * 1024  
ALLOWED_EXTENSIONS = {
    "csv": "csv",
    "xlsx": "xlsx",
    "xls": "xlsx",
    "txt": "txt",
    "vcf": "vcf"
}


class UploadService:
    """
    Industrial Asynchronous Bulk Ingestion Engine.
    Parses CSV, Excel, TXT, and native VCF files smoothly using direct stream matrix mappings.
    """

    @classmethod
    async def validate_and_stage_upload(cls, file: UploadFile, db: AsyncSession) -> Upload:
        file.file.seek(0, io.SEEK_END)
        actual_size = file.file.tell()
        file.file.seek(0)

        if actual_size > MAX_UPLOAD_SIZE:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="File is too large. Max capacity limit is 25MB."
            )

        filename = file.filename.lower()
        ext = filename.split(".")[-1] if "." in filename else ""
        
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail="Invalid file layout. System accepts only .csv, .xlsx, .txt, or .vcf formats."
            )

        system_safe_name = f"bulk_{uuid.uuid4().hex}.{ALLOWED_EXTENSIONS[ext]}"
        sanitized_original = re.sub(r"[^a-zA-Z0-9_\.\-]", "_", file.filename or "unknown")

        batch = Upload(
            filename=system_safe_name,
            original_filename=sanitized_original,
            file_size=actual_size,
            status="PENDING",
            total_records_found=0
        )
        db.add(batch)
        await db.commit()
        await db.refresh(batch)
        return batch

    @classmethod
    async def stream_and_ingest_bulk_file(cls, db: AsyncSession, upload_batch: Upload, file: UploadFile) -> int:
        upload_batch.status = "PROCESSING"
        await db.commit()

        file_contents = await file.read()
        filename = file.filename.lower()
        ext = filename.split(".")[-1] if "." in filename else ""
        detected_format = ALLOWED_EXTENSIONS.get(ext)

        parsed_records: List[Dict[str, str]] = []

        try:
            if detected_format == "csv":
                parsed_records = cls._stream_csv(file_contents)
            elif detected_format == "xlsx":
                parsed_records = cls._stream_xlsx(file_contents)
            elif detected_format == "txt":
                parsed_records = cls._stream_txt(file_contents)
            elif detected_format == "vcf":
                parsed_records = cls._stream_vcf(file_contents)

            if not parsed_records:
                raise ContactProcessingError("The uploaded document grid contains zero executable phone lines.")

            upload_batch.total_records_found = len(parsed_records)
            await db.commit()

            successful_inserts = 0
            
            for record in parsed_records:
                phone = record["phone_number"]
                
                if re.match(r"^0\d{9}$", phone):
                    phone = f"+255{phone[1:]}"
                elif re.match(r"^255\d{9}$", phone):
                    phone = f"+{phone}"
                elif not re.match(r"^\+255\d{9}$", phone):
                    continue  

                exist_query = select(Contact).where(and_(Contact.phone_number == phone, Contact.deleted_at == None))
                res = await db.execute(exist_query)
                if res.scalars().first():
                    continue

                new_row = Contact(
                    first_name=record["first_name"],
                    last_name=record["last_name"],
                    phone_number=phone,
                    consent_given=True,
                    is_approved=False,  
                    terms_version="v1.0"
                )
                db.add(new_row)
                successful_inserts += 1

            upload_batch.status = "COMPLETED"
            await db.commit()
            logger.info(f"Bulk ingestion finished safely. Staged total: {successful_inserts}")
            return successful_inserts

        except Exception as e:
            await db.rollback()
            upload_batch.status = "FAILED"
            await db.commit()
            logger.error(f"Batch execution failed: {str(e)}")
            raise ContactProcessingError(f"System processing failed: {str(e)}")

    # 1. FIXED CSV PARSER: Inasoma vichwa vya nguzo kwa uhakika
    @staticmethod
    def _stream_csv(contents: bytes) -> List[Dict[str, str]]:
        try:
            decoded = contents.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = contents.decode("latin-1")

        stream = io.StringIO(decoded)
        reader = csv.DictReader(stream)
        
        # Normalization ya vichwa vya nguzo
        if reader.fieldnames:
            reader.fieldnames = [f.strip().lower() for f in reader.fieldnames]

        records = []
        for row in reader:
            phone_raw = row.get("phone_number") or row.get("phone") or row.get("namba") or row.get("namba ya simu")
            full_name_raw = row.get("name") or row.get("first_name") or row.get("jina") or "Bulk"
            l_name_raw = row.get("last_name") or ""

            if phone_raw:
                clean_phone = re.sub(r"[\s\-\(\)\+]", "", str(phone_raw))
                if clean_phone:
                    name_parts = str(full_name_raw).strip().split()
                    f_name = name_parts[0] if name_parts else "Bulk"
                    # Kama alama ya jina la pili haipo kwenye CSV, unganisha yaliyobaki
                    l_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else str(l_name_raw).strip()
                    records.append({"first_name": f_name, "last_name": l_name, "phone_number": clean_phone})
        return records

    # 2. FIXED EXCEL PARSER: Inatafuta mkaazi wa seli kwa uhakika wa 100%
    @staticmethod
    def _stream_xlsx(contents: bytes) -> List[Dict[str, str]]:
        stream = io.BytesIO(contents)
        wb = load_workbook(stream, read_only=True, data_only=True)
        sheet = wb.active
        if not sheet:
            wb.close()
            return []

        records = []
        rows = list(sheet.iter_rows(values_only=True))
        wb.close()
        
        if not rows or len(rows) < 2:
            return []

        # FIXED HEADER LOOKUP: Tunachukua row ya kwanza tu (Index 0)
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        
        p_idx = -1
        f_idx = -1
        l_idx = -1
        
        for idx, h in enumerate(headers):
            if h in ["phone_number", "phone", "namba", "namba ya simu"]: p_idx = idx
            if h in ["first_name", "name", "jina"]: f_idx = idx
            if h in ["last_name"]: l_idx = idx

        # Fallback mifumo ya dharura kama faili haina headers rasmi
        if p_idx == -1: p_idx = 1 if len(headers) > 1 else 0
        if f_idx == -1: f_idx = 0

        for row in rows[1:]:
            if not row or len(row) <= p_idx or row[p_idx] is None:
                continue
                
            clean_phone = re.sub(r"[\s\-\(\)\+]", "", str(row[p_idx]))
            if not clean_phone:
                continue
                
            full_name_raw = str(row[f_idx]).strip() if (f_idx < len(row) and row[f_idx]) else "Bulk"
            l_name_raw = str(row[l_idx]).strip() if (l_idx != -1 and l_idx < len(row) and row[l_idx]) else ""
            
            name_parts = full_name_raw.split()
            f_name = name_parts[0] if name_parts else "Bulk"
            l_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else l_name_raw
            
            records.append({"first_name": f_name, "last_name": l_name, "phone_number": clean_phone})
        return records

    # 3. PARSER YA KUSOMA PLAIN TEXT FILES
    @staticmethod
    def _stream_txt(contents: bytes) -> List[Dict[str, str]]:
        try:
            decoded = contents.decode("utf-8")
        except UnicodeDecodeError:
            decoded = contents.decode("latin-1")

        records = []
        lines = decoded.splitlines()
        for line in lines:
            line_str = line.strip()
            if not line_str:
                continue
            phone_match = re.search(r"(\+?255\d{9}|0\d{9}|\d{9,12})", line_str)
            if phone_match:
                phone = re.sub(r"[\s\-\(\)\+]", "", phone_match.group(1))
                name_part = line_str.replace(phone_match.group(1), "").replace(",", "").replace("|", "").strip()
                name_parts = name_part.split()
                f_name = name_parts[0] if name_parts else "Bulk"
                l_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
                records.append({"first_name": f_name, "last_name": l_name, "phone_number": phone})
        return records

    # 4. FIXED VCF PARSER: Re.MULTILINE inakamata FN na TEL zote mfululizo
    @staticmethod
    def _stream_vcf(contents: bytes) -> List[Dict[str, str]]:
        try:
            vcf_content = contents.decode("utf-8", errors="ignore")
        except Exception:
            vcf_content = contents.decode("latin-1", errors="ignore")

        records = []
        cards = re.findall(r"BEGIN:VCARD.*?END:VCARD", vcf_content, re.DOTALL | re.IGNORECASE)
        
        for card in cards:
            # FIXED REGEX: re.MULTILINE inaruhusu kusoma katikati ya kila mstari kwenye kadi
            fn_match = re.search(r"^FN:(.*?)$", card, re.MULTILINE | re.IGNORECASE)
            n_match = re.search(r"^N:(.*?)$", card, re.MULTILINE | re.IGNORECASE)
            tel_match = re.search(r"^TEL;.*?[: ](.*?)$", card, re.MULTILINE | re.IGNORECASE)
            
            if not tel_match:
                continue
            
            phone = re.sub(r"[\s\-\(\)\+]", "", tel_match.group(1))
            if not phone:
                continue
            
            full_name = "Bulk Contact"
            if fn_match and fn_match.group(1).strip():
                full_name = fn_match.group(1).strip()
            elif n_match and n_match.group(1).strip():
                full_name = n_match.group(1).replace(";", " ").strip()
            
            name_parts = full_name.split()
            f_name = name_parts[0] if name_parts else "Bulk"
            l_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
            
            records.append({"first_name": f_name, "last_name": l_name, "phone_number": phone})
        return records
